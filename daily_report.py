import pyautogui
import time
import openpyxl
import pyperclip
import os
from datetime import datetime
pyautogui.FAILSAFE = True
pyautogui.PAUSE = 1
print("step 1: open the chrome browser")
time.sleep(2)
pyautogui.hotkey('win', 'r')
time.sleep(1)
pyautogui.typewrite('chrome')
time.sleep(1)
pyautogui.press('enter')
time.sleep(3)
print("step 2: go to the website")
pyautogui.hotkey('command', 't',interval=0.1)
time.sleep(1)
import subprocess

chrome_path = r"C:\Program Files\Google\Chrome\Application\chrome.exe"

subprocess.Popen([
    chrome_path,
    "--profile-directory=Profile 2"
])
pyautogui.write('https://www.accuweather.com/en/in/chennai/206671/weather-forecast/206671', interval=0.1)

time.sleep(1)
pyautogui.press('enter')
time.sleep(5)
import pyautogui
from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By

# ---------------------------------
# STEP 1: Fetch data from website
# ---------------------------------

print("Opening weather website...")

driver = webdriver.Chrome()

driver.get(
    "https://www.accuweather.com/en/in/chennai/206671/weather-forecast/206671"
)

time.sleep(5)

page_text = driver.find_element(By.TAG_NAME, "body").text

# Example: search for a line containing temperature symbol
fetched_data = "Weather data not found"

for line in page_text.split("\n"):
    if "°" in line:
        fetched_data = line
        break


print("Fetched Data:", fetched_data)

driver.quit()
# ---------------------------------
# STEP 2: Prepare file names
# ---------------------------------

today = datetime.now().strftime("%Y-%m-%d")

excel_file = f"daily_report_{today}.xlsx"
screenshot_file = f"daily_report_{today}.png"
# ---------------------------------
# STEP 3: Create/Open Excel file
# ---------------------------------

if os.path.exists(excel_file):
    wb = load_workbook(excel_file)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active

    ws["A1"] = "Date & Time"
    ws["B1"] = "Fetched Data"
    ws["C1"] = "Comment"
# ---------------------------------
# STEP 4: Add new row
# ---------------------------------

current_time = datetime.now().strftime("%d-%m-%Y %H:%M:%S")

comment = "Good for outdoor activities"

next_row = ws.max_row + 1

ws.cell(next_row, 1, current_time)
ws.cell(next_row, 2, fetched_data)
ws.cell(next_row, 3, comment)

# ---------------------------------
# STEP 5: Save Excel
# ---------------------------------

wb.save(excel_file)

print("Excel saved:", excel_file)

# ---------------------------------
# STEP 6: Open Excel
# ---------------------------------

os.startfile(excel_file)

time.sleep(5)
# ---------------------------------
# STEP 7: Take Screenshot
# ---------------------------------

screenshot = pyautogui.screenshot()
screenshot.save(screenshot_file)

print("Screenshot saved:", screenshot_file)

print("Task Completed Successfully")
